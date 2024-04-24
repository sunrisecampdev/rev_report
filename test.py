import openpyxl

# DONOR_ID = "Donor ID"
# SPOUSE_FIRST = "Household: Spouse First Name"
# SPOUSE_LAST = "Household: Spouse Last Name"
# PRIMARY_FIRST = "Household: Primary First Name"
# PRIMARY_LAST = "Household: Primary Last Name"
# COMPANY = "Company Name"
# EVENT_STATUS = "Event Status"
# ATTENDEES = "Attendees"
# DATE = "Date"
# SALE_TYPE = "Type"
# AMOUNT = "Amount"
# ATTRIBUTE = "Connection"

class revenueReport:
    def __init__(self, workbook, sheet1, sheet2):
        self.workbook = workbook
        self.sheet1 = sheet1
        self.sheet2 = sheet2
        self.HEADER_ROW = self.getHeaderRowIndex()-1
        self.MAX_COL = self.sheet1.max_column
        self.colIndex = 1
        self.headerDict = {
            "DONOR_ID" : "Donor ID",
            "SPOUSE_FIRST" : "Household: Spouse First Name",
            "SPOUSE_LAST" : "Household: Spouse Last Name",
            "PRIMARY_FIRST" : "Household: Primary First Name",
            "PRIMARY_LAST" : "Household: Primary Last Name",
            "COMPANY" : "Company Name",
            "EVENT_STATUS" : "Event Status",
            "ATTENDEES" : "Attendees",
            "DATE" : "Date",
            "SALE_TYPE" : "Type",
            "AMOUNT" : "Amount",
            "CONNECTION" : "Attribute"
        }
        self.headerOrder = ["DONOR_ID",
                            "SPOUSE_FIRST",
                            "SPOUSE_LAST",
                            "PRIMARY_FIRST",
                            "PRIMARY_LAST",
                            "COMPANY",
                            "EVENT_STATUS",
                            "ATTENDEES",
                            "DATE",
                            "SALE_TYPE",
                            "AMOUNT",
                            "CONNECTION"
                            ]

    def incColIndex(self):
        self.colIndex += 1

    def getWorkbook(self):
        return self.workbook
        
    def getHeaderRowIndex(self):
        """Returns the row index of the header row"""
        counter = 0
        for colCell in self.sheet1['A']:
            counter += 1
            if colCell.value == "Item":
                return counter
        
    def getColIndex(self, colName, headerRow, MAX_COL):
        for headerColIndex in range(1, MAX_COL):
            if self.sheet1.cell(row=headerRow, column=headerColIndex).value == colName:
                return headerColIndex
            
    def getColValues(self, targetCol):
        """Returns a list containing a tuple of all values from the respective column"""
        valueList = []
        for column in self.sheet1.iter_cols():
            colName = column[self.HEADER_ROW].value
            if colName == targetCol:
                for cell in column:
                    if cell.row > self.HEADER_ROW:
                        valueList.append(cell.value)
        return valueList

    def transferCol(self, colList, headerName):
        rowIndex = 1
        for colCellValue in colList:
            currentCell = self.sheet2.cell(row=rowIndex, column=self.colIndex)
            currentCell.value = colCellValue
            self.cellFormat(currentCell, headerName)
            rowIndex += 1
        self.incColIndex()
        return
    
    def cellFormat(self, cell, headerName):
        if headerName == "DATE":
            cell.number_format = "mm/dd/yyyy"
        if headerName == "AMOUNT":
            cell.number_format = "$#,##0.00"

    def checkCustomCol(self):
        return

    def transferFirstName(self):
        return

    def transferLastName(self):
        return
    
    def transferCompany(self):
        return
    
    def transferAllCols(self):
        for header in self.headerOrder:
            currentCol = self.getColValues(self.headerDict[header])
            self.transferCol(currentCol, header)
        return
        
wb = openpyxl.load_workbook('newrev.xlsx')
ws1 = wb['Sheet1']
ws2 = wb.create_sheet("Sheet2")

newReport = revenueReport(wb, ws1, ws2)

# newReport.transferCol(newReport.getColValues(newReport.headerDict["DONOR_ID"]))

newReport.transferAllCols()

# for x in range(1, 20):
#     print(newReport.sheet2.cell(row=x, column=1).value)

newbook = newReport.getWorkbook()
newbook.save("superReport.xlsx")


















